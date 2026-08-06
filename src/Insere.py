from openpyxl import load_workbook
import os
from openpyxl.styles import Alignment, PatternFill
    
def dia(idd,periodo, linha_busca,tipo,hor,wb,aba,Dia_hoje,nome_arquivo):
    
    linha_registro = linha_busca + 1

    # Estilo de alinhamento centralizado reutilizável
    alinhamento_centro = Alignment(horizontal="center", vertical="center")

    # 1. Coluna 1 (A): Dia
    dia = aba.cell(row=linha_registro, column=1, value=Dia_hoje)
    dia.alignment = alinhamento_centro
    dia.fill = PatternFill(
        start_color="B5EDBA", end_color="B5EDBA", fill_type="solid"
    )

    # 2. Coluna 2 (B): ID
    iid = aba.cell(row=linha_registro, column=2, value=int(idd))
    iid.alignment = alinhamento_centro

    # 3. Coluna 3 (C): Período
    per = aba.cell(row=linha_registro, column=3, value=periodo)
    per.alignment = alinhamento_centro

    # 4. Coluna 4 (D): Tempo / Hora
    tem = aba.cell(row=linha_registro, column=4, value=int(hor))
    tem.alignment = alinhamento_centro

    # 5. Coluna 5 (E): Situação / Tipo
    sit = aba.cell(row=linha_registro, column=5, value=tipo)
    sit.alignment = alinhamento_centro

    if tipo == "Presencial":
        sit.fill= PatternFill(
        start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
    )
    else:
        sit.fill= PatternFill(
                start_color="FFFFC5", end_color="FFFFC5", fill_type="solid"
            )
    # Salva as alterações no arquivo
    wb.save(nome_arquivo)







    