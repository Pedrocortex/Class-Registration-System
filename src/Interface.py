import tkinter as tk
from tkinter import messagebox
from Insere import dia
import os
from openpyxl import load_workbook
from datetime import datetime
from tkinter import ttk


switcher = False
Dia_hoje = datetime.now().strftime("%d/%m/%Y")

def atualiza():

    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    nome_arquivo = os.path.join(diretorio_atual, "Pasta2.xlsx")
    wb = load_workbook(nome_arquivo)
    aba = wb.active

    return wb,aba,nome_arquivo, 


def enviar_dados(periodo,idd, tipo,hor,janela):    
    wb,aba,nome_arquivo = atualiza()
    linha_busca = wb.active.max_row 
    if not idd:
        messagebox.showwarning("Aviso", "Por favor, digite o o id do aluno!")
        return

    dia(idd,periodo, linha_busca,tipo,hor,wb,aba,Dia_hoje,nome_arquivo)

    janela.destroy()
    switcher=True
    recebe_dados(periodo,tipo )

    return switcher


# Cria a janela principal
def recebe_dados(p,s):
        
    janela = tk.Tk()
    janela.title("Portal de Cadastro")
    janela.geometry("300x410") # Define o tamanho da janela (Largura x Altura)
    
    btn_atualizar = tk.Button(
        janela,
        text="Apagar número",
        command=lambda: (janela.destroy(), recebe_dados(p, s))

)
    btn_atualizar.pack(pady=5)

    label = tk.Label(janela, text="Escolha o horário:")
    label.pack(pady=10)

    # 1. Criar a caixa de seleção (Combobox) com os valores de 1 a 12
    horario= [  int(i) for i in range(1, 13)]
    hor = ttk.Combobox(janela, values=horario, state="readonly")
    hor.pack(pady=5)

    # Definir um valor padrão inicial (opcional)
    hor.current(0)  

    # 2. Campo: ID
    lbl_id = tk.Label(janela, text="ID do aluno")
    lbl_id.pack(pady=5)
    
    entrada_id = tk.Entry(janela)
    entrada_id.pack(pady=5)

    # 3. Campo: Período (Usando botões de rádio para PM e AM)
    lbl_periodo = tk.Label(janela, text="Período:")
    lbl_periodo.pack(pady=5)

    periodo = tk.StringVar(value=p) # Define "PM" como padrão inicial

    tk.Radiobutton(janela, text="AM", variable=periodo, value="AM").pack()
    tk.Radiobutton(janela, text="PM", variable=periodo, value="PM").pack()

    # 4. Campo: Status (Usando botões de rádio para PM  e AM)
    lbl_situacao = tk.Label(janela, text="Tipo de atendimento:")
    lbl_situacao.pack(pady=5)

    tipo = tk.StringVar(value=s) 

    tk.Radiobutton(janela, text="Presencial", variable=tipo  , value="Presencial").pack()
    tk.Radiobutton(janela, text="Online", variable=tipo, value="Online").pack()

    
    # 4. Botão de Enviar
    btn_enviar = tk.Button(
        janela,
        text="Gravar Dados",
        command=lambda: enviar_dados(
        periodo.get(),
        entrada_id.get(),
        tipo.get(),
        hor.get(),
        janela
        ),
        bg="#4A90E2",
        fg="white"
    )

    btn_enviar.pack(pady=20)

    janela.mainloop()
    
if not switcher:
    recebe_dados ("PM","Presencial")
